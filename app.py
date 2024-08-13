from flask import Flask, request, jsonify, render_template
import pandas as pd
from openpyxl import load_workbook
import os

app = Flask(__name__, static_folder='static', template_folder='templates')

# Path to the Excel file
EXCEL_FILE = 'attendance.xlsx'

# Ensure the Excel file exists with the correct headers
if not os.path.exists(EXCEL_FILE):
    df = pd.DataFrame(columns=['USN', 'Student Name'])
    df.to_excel(EXCEL_FILE, index=False)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/save_attendance', methods=['POST'])
def save_attendance():
    data = request.json
    date = data.get('date')
    if not date:
        return jsonify({'error': 'Date is required'}), 400

    # Load existing data or initialize new DataFrame
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        if 'Attendance' in wb.sheetnames:
            sheet = wb['Attendance']
            df = pd.read_excel(EXCEL_FILE, sheet_name='Attendance')
        else:
            sheet = wb.create_sheet('Attendance')
            df = pd.DataFrame(columns=['USN', 'Student Name'])
    else:
        wb = load_workbook(EXCEL_FILE)
        df = pd.DataFrame(columns=['USN', 'Student Name'])
        sheet = wb.create_sheet('Attendance')

    # Prepare the DataFrame for the new data
    attendance_data = pd.DataFrame.from_dict(data, orient='index').reset_index()
    attendance_data.columns = ['Student Name', date]

    # Update or append new data
    if date in df.columns:
        df.update(attendance_data.set_index('Student Name'))
    else:
        df = pd.merge(df, attendance_data, on='Student Name', how='left')

    # Save to Excel
    wb.remove(sheet)
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Attendance', index=False)

    return jsonify({'message': 'Attendance saved successfully!'}), 200

if __name__ == '__main__':
    app.run(debug=True)
