from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook
import os

app = Flask(__name__)

# Define the path to the workbook, ensuring it works on Render
workbook_path = os.path.join(os.path.dirname(__file__), 'data', 'attendance.xlsx')
workbook = load_workbook(workbook_path)
sheet = workbook.active

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/save_attendance', methods=['POST'])
def save_attendance():
    data = request.json
    date_day = data.pop('date')  # Extract date/day from the JSON data

    # Check if the date/day already exists in the sheet
    date_column = None
    for col in range(3, sheet.max_column + 1):  # Columns 1 and 2 are fixed for USN and Student Name
        if sheet.cell(row=1, column=col).value == date_day:
            date_column = col
            break

    if date_column is None:
        # If the date/day does not exist, create a new column
        date_column = sheet.max_column + 1
        sheet.cell(row=1, column=date_column).value = date_day

    # Iterate over the attendance data to update the Excel sheet
    for student, attendance in data.items():
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=2).value == student:
                sheet.cell(row=row, column=date_column).value = attendance
                break

    # Save the updated workbook
    workbook.save(workbook_path)

    return jsonify({'message': 'Attendance saved successfully!'})

if __name__ == '__main__':
    # Ensure the app runs on the correct host and port for Render
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 10000)))
